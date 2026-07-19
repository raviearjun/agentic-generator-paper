#!/usr/bin/env python3
"""
generate_xlsx.py

Runs the core SHACL shapes graph (core.shapes.ttl) against the generated KG
corpus and writes an .xlsx summary in the same visual style as
script_to_kg/evaluation/evaluate_kgs.py's write_xlsx() (same fills/fonts,
same "Per-File" + "Group Summary" sheet layout convention).

Scope matches the SHACL subsection of evaluation.tex: the P5/GPT-5-mini
corpus (50 files), minus six examples with no genuine agentic implementation
in their source (three LangGraph, three Mastra AI — see shacl/README.md
"Removed projects" and evaluation.tex Cref{sec:eval-shacl} for the
justification), leaving 44 files. Files that fail to parse as Turtle are
still listed (status "Parse Error") but excluded from the conformance
counts/averages, matching how the paper reports 44 -> 41.

Usage:
    python3 generate_xlsx.py
Requires pyshacl (pip install pyshacl) and openpyxl (already a
script_to_kg/evaluation dependency).
"""

import re
import subprocess
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required: pip install openpyxl")
    sys.exit(1)

HERE = Path(__file__).resolve().parent
REPO = HERE.parents[2]
GENERATED_KGS_DIR = REPO / "script_to_kg" / "generated_kgs"
CORE_SHAPES = HERE / "core.shapes.ttl"
OUT_XLSX = HERE / "shacl_evaluation_summary.xlsx"

# Six examples excluded from the SHACL corpus (no genuine agentic
# implementation in their source) — see evaluation.tex, section eval-shacl.
EXCLUDED = {
    "open-code", "stockbroker", "trip-planner",       # LangGraph
    "a2a", "agent1", "mcp-registry-registry",          # Mastra AI
}

# ── same visual style as evaluate_kgs.py ────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
GOOD_FILL    = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL    = PatternFill("solid", fgColor="FFEB9C")
BAD_FILL     = PatternFill("solid", fgColor="FFC7CE")
WHITE_FONT   = Font(color="FFFFFF", bold=True, name="Calibri")
BOLD_FONT    = Font(bold=True, name="Calibri")
REG_FONT     = Font(name="Calibri")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _set_cell(ws, row, col, value, fill=None, font=None, align="center", border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    cell.font = font or REG_FONT
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        cell.border = THIN_BORDER


def _status_fill(status):
    if status == "Conformant":
        return GOOD_FILL
    if status == "Violations":
        return BAD_FILL
    if status == "Warnings only":
        return WARN_FILL
    return None  # Parse Error / Excluded — no fill


# ── run pyshacl over the corpus ─────────────────────────────────────────────

def run_shacl(data_path: Path):
    """Return (parsed_ok, conforms, n_violations, n_warnings, violation_props, warning_props)."""
    r = subprocess.run(
        ["pyshacl", "-s", str(CORE_SHAPES), str(data_path)],
        capture_output=True, text=True,
    )
    out = r.stdout
    if not out:
        return False, None, None, None, [], []

    conforms = "Conforms: True" in out
    n_viol = out.count("Severity: sh:Violation")
    n_warn = out.count("Severity: sh:Warning")

    viol_props, warn_props = [], []
    current_sev = None
    for line in out.split("\n"):
        s = line.strip()
        if s.startswith("Severity:"):
            current_sev = s.split(":", 1)[1].strip()
        elif s.startswith("Message:") and current_sev:
            msg = s.split(":", 1)[1].strip()
            m = re.search(r":(\w+)", msg)
            prop = m.group(1) if m else "?"
            if current_sev == "sh:Violation":
                viol_props.append(prop)
            elif current_sev == "sh:Warning":
                warn_props.append(prop)
            current_sev = None

    return True, conforms, n_viol, n_warn, viol_props, warn_props


def collect_records():
    records = []
    for fw_dir in sorted(GENERATED_KGS_DIR.iterdir()):
        if not fw_dir.is_dir():
            continue
        framework = fw_dir.name
        for f in sorted(fw_dir.glob("*_instances.ttl")):
            stem = f.name[: -len("_instances.ttl")]
            excluded = stem in EXCLUDED
            rec = {
                "framework": framework, "file": f.name, "stem": stem,
                "excluded": excluded,
            }
            if excluded:
                rec.update(parsed=None, conforms=None, n_viol=None, n_warn=None,
                           viol_props=[], warn_props=[])
            else:
                parsed, conforms, n_viol, n_warn, viol_props, warn_props = run_shacl(f)
                rec.update(parsed=parsed, conforms=conforms, n_viol=n_viol, n_warn=n_warn,
                           viol_props=viol_props, warn_props=warn_props)
            records.append(rec)
            if excluded:
                note = "excluded"
            elif not rec["parsed"]:
                note = "parse error"
            elif rec["conforms"]:
                note = "conforms"
            else:
                note = f"{rec['n_viol']}V/{rec['n_warn']}W"
            print(f"  [{framework}] {f.name}: {note}")
    return records


def row_status(rec):
    if rec["excluded"]:
        return "Excluded"
    if not rec["parsed"]:
        return "Parse Error"
    if rec["conforms"]:
        return "Conformant"
    if rec["n_viol"] and rec["n_viol"] > 0:
        return "Violations"
    return "Warnings only"


# ── xlsx writer ──────────────────────────────────────────────────────────

def write_xlsx(records, path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Per-File Results ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Per-File Results"
    ws1.freeze_panes = "C2"

    headers = ["Framework", "File", "Status", "Conforms",
               "Violations", "Warnings", "Top Violation Props", "Top Warning Props"]
    for col, h in enumerate(headers, 1):
        _set_cell(ws1, 1, col, h, fill=HEADER_FILL, font=WHITE_FONT)
    ws1.row_dimensions[1].height = 22

    row_idx = 2
    for rec in records:
        status = row_status(rec)
        alt = ALT_FILL if row_idx % 2 == 0 else None
        fill_status = _status_fill(status) or alt

        viol_summary = ", ".join(f"{p}({c})" for p, c in
                                  sorted(_count(rec["viol_props"]).items(), key=lambda x: -x[1]))
        warn_summary = ", ".join(f"{p}({c})" for p, c in
                                  sorted(_count(rec["warn_props"]).items(), key=lambda x: -x[1]))

        vals = [
            rec["framework"], rec["file"], status,
            "" if rec["conforms"] is None else ("✓" if rec["conforms"] else "✗"),
            rec["n_viol"] if rec["n_viol"] is not None else "",
            rec["n_warn"] if rec["n_warn"] is not None else "",
            viol_summary, warn_summary,
        ]
        for col, v in enumerate(vals, 1):
            fill = fill_status if col in (3, 4) else alt
            _set_cell(ws1, row_idx, col, v, fill=fill, font=REG_FONT,
                      align="left" if col in (1, 2, 7, 8) else "center")
        row_idx += 1

    col_widths = [12, 42, 14, 10, 11, 10, 40, 40]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Group Summary (per framework, matches evaluation.tex table) ──
    ws2 = wb.create_sheet("Group Summary")
    ws2.freeze_panes = "A2"

    headers2 = ["Framework", "Files Evaluated", "Excluded (non-agentic)",
                "Parse Errors", "Conformant", "Violations", "Warnings",
                "Avg Violations/File", "Avg Warnings/File"]
    for col, h in enumerate(headers2, 1):
        _set_cell(ws2, 1, col, h, fill=HEADER_FILL, font=WHITE_FONT)
    ws2.row_dimensions[1].height = 30

    by_fw = defaultdict(list)
    for rec in records:
        by_fw[rec["framework"]].append(rec)

    row_idx = 2
    totals = defaultdict(int)
    for framework in sorted(by_fw):
        recs = by_fw[framework]
        excluded_n = sum(1 for r in recs if r["excluded"])
        parse_err_n = sum(1 for r in recs if not r["excluded"] and not r["parsed"])
        evaluated = [r for r in recs if not r["excluded"] and r["parsed"]]
        n_eval = len(evaluated)
        conformant = sum(1 for r in evaluated if r["conforms"])
        n_viol = sum(r["n_viol"] for r in evaluated)
        n_warn = sum(r["n_warn"] for r in evaluated)
        avg_v = n_viol / n_eval if n_eval else 0.0
        avg_w = n_warn / n_eval if n_eval else 0.0

        totals["evaluated"] += n_eval
        totals["excluded"] += excluded_n
        totals["parse_err"] += parse_err_n
        totals["conformant"] += conformant
        totals["viol"] += n_viol
        totals["warn"] += n_warn

        alt = ALT_FILL if row_idx % 2 == 0 else None
        vals = [framework, n_eval, excluded_n, parse_err_n, conformant,
                n_viol, n_warn, round(avg_v, 2), round(avg_w, 2)]
        for col, v in enumerate(vals, 1):
            _set_cell(ws2, row_idx, col, v, fill=alt, font=REG_FONT,
                      align="left" if col == 1 else "center")
        row_idx += 1

    avg_v_total = totals["viol"] / totals["evaluated"] if totals["evaluated"] else 0.0
    avg_w_total = totals["warn"] / totals["evaluated"] if totals["evaluated"] else 0.0
    vals = ["Total", totals["evaluated"], totals["excluded"], totals["parse_err"],
            totals["conformant"], totals["viol"], totals["warn"],
            round(avg_v_total, 2), round(avg_w_total, 2)]
    for col, v in enumerate(vals, 1):
        _set_cell(ws2, row_idx, col, v, fill=SUBHEAD_FILL, font=WHITE_FONT,
                  align="left" if col == 1 else "center")

    col_widths2 = [12, 15, 20, 13, 12, 11, 11, 18, 18]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Violation & Warning Types ───────────────────────────────
    ws3 = wb.create_sheet("Violation & Warning Types")
    ws3.freeze_panes = "A2"

    viol_counter = defaultdict(int)
    warn_counter = defaultdict(int)
    for rec in records:
        for p in rec["viol_props"]:
            viol_counter[p] += 1
        for p in rec["warn_props"]:
            warn_counter[p] += 1

    _set_cell(ws3, 1, 1, "Violations (sh:Violation) by property", fill=HEADER_FILL, font=WHITE_FONT)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    _set_cell(ws3, 2, 1, "Property", fill=SUBHEAD_FILL, font=WHITE_FONT)
    _set_cell(ws3, 2, 2, "Count", fill=SUBHEAD_FILL, font=WHITE_FONT)
    r = 3
    for prop, cnt in sorted(viol_counter.items(), key=lambda x: -x[1]):
        alt = ALT_FILL if r % 2 == 1 else None
        _set_cell(ws3, r, 1, prop, fill=alt, align="left")
        _set_cell(ws3, r, 2, cnt, fill=alt)
        r += 1

    r += 1
    _set_cell(ws3, r, 1, "Warnings (sh:Warning) by property", fill=HEADER_FILL, font=WHITE_FONT)
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    _set_cell(ws3, r, 1, "Property", fill=SUBHEAD_FILL, font=WHITE_FONT)
    _set_cell(ws3, r, 2, "Count", fill=SUBHEAD_FILL, font=WHITE_FONT)
    r += 1
    for prop, cnt in sorted(warn_counter.items(), key=lambda x: -x[1]):
        alt = ALT_FILL if r % 2 == 1 else None
        _set_cell(ws3, r, 1, prop, fill=alt, align="left")
        _set_cell(ws3, r, 2, cnt, fill=alt)
        r += 1

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 12

    wb.save(path)
    print(f"wrote {path.relative_to(REPO)}")


def _count(items):
    d = defaultdict(int)
    for x in items:
        d[x] += 1
    return d


def main():
    print("Running SHACL validation over the corpus...")
    records = collect_records()
    write_xlsx(records, OUT_XLSX)


if __name__ == "__main__":
    main()
